import streamlit as st
import pandas as pd
import os
import json
from datetime import datetime
import plotly.express as px
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="SOX Control Monitoring Platform", layout="wide")

# ---------------- FILE PATHS ---------------- #

UPLOAD_DIR = "data/uploads"
OUTPUT_DIR = "data/output"
LOG_FILE = "data/upload_log.csv"
CHANGE_FILE = "data/change_analysis.csv"
LATEST_FILE = "data/latest_data.csv"
STATE_FILE = "data/app_state.json"

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------- HEADER ---------------- #

st.markdown("""
<div style="background:linear-gradient(90deg,#0f172a,#1e3a8a);
padding:30px;border-radius:12px;color:white;margin-bottom:20px;">
<h1>SOX Control Monitoring Platform</h1>
<p>Internal Audit Analytics Dashboard</p>
</div>
""", unsafe_allow_html=True)

# ---------------- SIDEBAR ---------------- #

st.sidebar.title("Navigation")

page = st.sidebar.radio(
    "Select Page",
    ["Executive Dashboard","Change Analysis","Upload History","Raw Data"]
)

uploaded_file = st.sidebar.file_uploader("Upload SOX Dashboard", type=["xlsx"])

# ---------------- VALIDATION ---------------- #

def validate_dataset(df):
    required = [
        "PROCESS_UID","CYCLE","Test Name",
        "TESTS__TEST_SECTION","TESTS__STATUS",
        "TESTS__EFFECTIVENESS","TESTS__TESTER_USER",
        "TESTS__REVIEWER_USER","TESTS__SECONDARY_REVIEWER_USER",
        "TESTS__START_DATE","TESTS__END_DATE",
        "TESTS__DUE_DATE","PWC Reliance","Audit Team"
    ]

    missing=[c for c in required if c not in df.columns]

    if missing:
        st.error(f"Missing columns: {missing}")
        st.stop()

# ---------------- SAVE FILE ---------------- #

def save_file(uploaded_file):
    timestamp=datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename=f"{timestamp}.xlsx"
    path=os.path.join(UPLOAD_DIR,filename)

    with open(path,"wb") as f:
        f.write(uploaded_file.getbuffer())

    return filename,path

# ---------------- LOAD FILES ---------------- #

def load_versions():
    files=sorted(os.listdir(UPLOAD_DIR))
    files=[f for f in files if f.endswith(".xlsx")]

    if len(files)==0:
        return None,None

    latest=os.path.join(UPLOAD_DIR,files[-1])
    latest_df=pd.read_excel(latest,sheet_name="IA data")

    prev_df=None

    if len(files)>1:
        prev=os.path.join(UPLOAD_DIR,files[-2])
        prev_df=pd.read_excel(prev,sheet_name="IA data")

    return latest_df,prev_df

# ---------------- COMPARE ---------------- #

def compare_versions(old_df,new_df):

    key="Test Name"

    old_df[key]=old_df[key].astype(str).str.strip()
    new_df[key]=new_df[key].astype(str).str.strip()

    old_df=old_df.set_index(key)
    new_df=new_df.set_index(key)

    fields=list(new_df.columns)

    changes=[]

    common=old_df.index.intersection(new_df.index)

    for test in common:
        for field in fields:

            old=str(old_df.loc[test,field])
            new=str(new_df.loc[test,field])

            if old!=new:
                changes.append({
                    "Test Name":test,
                    "Field Changed":field,
                    "Old Value":old,
                    "New Value":new
                })

    return pd.DataFrame(changes)

# ---------------- UPLOAD LOG ---------------- #

def update_upload_log(filename, df):

    entry = {
        "File": filename,
        "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Rows": len(df)
    }

    if os.path.exists(LOG_FILE):
        log = pd.read_csv(LOG_FILE)
        log = pd.concat([log, pd.DataFrame([entry])], ignore_index=True)
    else:
        log = pd.DataFrame([entry])

    log.to_csv(LOG_FILE, index=False)

# ---------------- HIGHLIGHT FILE ---------------- #

def generate_highlight_file(changes):

    files=sorted(os.listdir(UPLOAD_DIR))
    files=[f for f in files if f.endswith(".xlsx")]

    latest=os.path.join(UPLOAD_DIR,files[-1])
    output=os.path.join(OUTPUT_DIR,"highlighted_changes.xlsx")

    wb=load_workbook(latest)
    ws=wb["IA data"]

    yellow=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")

    headers=[cell.value for cell in ws[1]]

    for _,row in changes.iterrows():

        test=row["Test Name"]
        field=row["Field Changed"]

        if field in headers:

            col=headers.index(field)+1

            for r in range(2,ws.max_row+1):
                if ws.cell(r,headers.index("Test Name")+1).value==test:
                    ws.cell(r,col).fill=yellow

    wb.save(output)

    return output

# ---------------- HANDLE UPLOAD ---------------- #

if uploaded_file:

    filename,path=save_file(uploaded_file)

    df=pd.read_excel(path,sheet_name="IA data")
    validate_dataset(df)

    update_upload_log(filename, df)

    latest_df,prev_df=load_versions()

    if latest_df is not None:
        latest_df.to_csv(LATEST_FILE,index=False)

    if latest_df is not None and prev_df is not None:

        changes=compare_versions(prev_df,latest_df)

        if not changes.empty:

            changes.to_csv(CHANGE_FILE,index=False)

            state = {
                "last_updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "rows": len(changes)
            }

            with open(STATE_FILE,"w") as f:
                json.dump(state,f)

# ---------------- LOAD DATA ---------------- #

latest_df=None
changes=None
state_info=None

if os.path.exists(LATEST_FILE):
    try:
        latest_df=pd.read_csv(LATEST_FILE)
    except:
        latest_df=None

if latest_df is None:
    latest_df,_=load_versions()

if os.path.exists(CHANGE_FILE):
    try:
        changes=pd.read_csv(CHANGE_FILE)
        if changes.empty:
            changes=None
    except:
        changes=None

if os.path.exists(STATE_FILE):
    try:
        with open(STATE_FILE,"r") as f:
            state_info=json.load(f)
    except:
        state_info=None

# ---------------- EXECUTIVE DASHBOARD ---------------- #

if page=="Executive Dashboard":

    if latest_df is not None:

        col1,col2=st.columns(2)

        total=len(latest_df)

        open_tests=latest_df[
            latest_df["TESTS__STATUS"].astype(str)
            .str.lower()
            .str.contains("open",na=False)
        ].shape[0]

        col1.metric("Total Tests",total)
        col2.metric("Open Tests",open_tests)

        status_counts=latest_df["TESTS__STATUS"].value_counts()

        c1,c2=st.columns(2)

        with c1:
            fig=px.pie(latest_df,names="TESTS__STATUS",hole=0.45)
            st.plotly_chart(fig,use_container_width=True)

        with c2:
            fig2=px.bar(status_counts)
            st.plotly_chart(fig2,use_container_width=True)

# ---------------- CHANGE ANALYSIS ---------------- #

elif page=="Change Analysis":

    st.subheader("Changes Since Last Upload")

    if state_info:
        st.success(f"Last Analysis: {state_info['last_updated']} | Changes: {state_info['rows']}")

    if changes is None:

        st.info("Upload at least two dashboards to detect changes")

    else:

        col1,col2=st.columns(2)

        test_filter=col1.multiselect("Test Name",changes["Test Name"].unique())
        field_filter=col2.multiselect("Field Changed",changes["Field Changed"].unique())

        filtered=changes.copy()

        if test_filter:
            filtered=filtered[filtered["Test Name"].isin(test_filter)]

        if field_filter:
            filtered=filtered[filtered["Field Changed"].isin(field_filter)]

        status_changes=len(filtered[filtered["Field Changed"]=="TESTS__STATUS"])
        tests_affected=filtered["Test Name"].nunique()
        fields_changed=filtered["Field Changed"].nunique()

        k1,k2,k3=st.columns(3)

        k1.metric("Status Changes",status_changes)
        k2.metric("Unique Tests Affected",tests_affected)
        k3.metric("Fields With Changes",fields_changed)

        st.dataframe(filtered,use_container_width=True)

        fig=px.bar(filtered["Field Changed"].value_counts())
        st.plotly_chart(fig,use_container_width=True)

        if st.button("Download Highlighted Excel"):

            path=generate_highlight_file(changes)

            with open(path,"rb") as f:
                st.download_button("Download Excel",f,
                                   file_name="highlighted_changes.xlsx")

# ---------------- HISTORY ---------------- #

elif page=="Upload History":

    if os.path.exists(LOG_FILE):
        log=pd.read_csv(LOG_FILE)
        st.dataframe(log,use_container_width=True)
    else:
        st.info("No uploads yet")

# ---------------- RAW DATA ---------------- #

elif page=="Raw Data":

    if latest_df is not None:
        st.dataframe(latest_df)
import os

if os.path.exists("data/latest_data.csv"):
    with open("data/latest_data.csv", "rb") as f:
        st.download_button("Download Latest Data", f, "latest_data.csv")

if os.path.exists("data/change_analysis.csv"):
    with open("data/change_analysis.csv", "rb") as f:
        st.download_button("Download Change Analysis", f, "change_analysis.csv")
